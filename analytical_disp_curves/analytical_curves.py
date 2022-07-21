import openpyxl

import numpy as np

import matplotlib
# matplotlib.use("pgf")
# matplotlib.rcParams.update({
#     "pgf.texsystem": "pdflatex",
#     'font.family': 'serif',
#     'text.usetex': True,
#     'pgf.rcfonts': False,
# })

import matplotlib.pyplot as plt

from pathlib import Path

"""
contains the most used functions for working with analytical
data from loading data to processing it

created by: Max Schmitz on 08/30/2021
"""


def load_analytical_data(
        material: str,
        analytical_path: str
) -> tuple:
    """
    load analytical obtained dispersion data from .npy files as well as
    mode names from .csv files

    returns:
        - ka: loaded wave number array
        - fa: loaded frequency array
        - mn: name of the respective modes

    args:
        - material: string which specifies the name of the
        - analytical_path: string or path object in which the analytical
                dispersion data is stored
    """

    # analytical_path = Path(__file__).parent.resolve()

    with open(analytical_path / Path(material + '_dispersion_data_analytically_f_.npy'), 'rb') as f:
        fa = np.load(f)
    with open(analytical_path / Path(material + '_dispersion_data_analytically_k_.npy'), 'rb') as f:
        ka = np.load(f)
    with open(analytical_path / Path(material + '_dispersion_data_analytically_mode_names_.csv'), 'rb') as f:
        mn = f.readlines()

    mn = [str(i)[2:4] for i in mn]

    return ka, fa, mn


def save_dispersion_np_array(
        output_name: str,
        data: float,
        fk: str,
        path: str = None
) -> None:
    """
    save numpy array as a binary file in an .npy file

    args:
        - output_name (str): name of the output file (usually name of material)
        - data (float): variable which should be stored. Either f or k
        - fk (str): either f or k - specify which of both was stored
        - path (str or Path object): specify where file should be stored
            if not in current folder of calling script
    """
    if path is not None:
        output_name = str(path / output_name)
    with open(Path(output_name + f'_analytically_{fk}_.npy'), 'wb') as file:
        np.save(file, data)


def save_mode_names(
        output_name: str,
        mode_names: list,
        path: str = None
) -> None:
    """
    save mode names of dispersion data in a list

    args:
        - output_name (str): name of the output file (should be name of material)
        - mode_names (list): list of strings with mode names
        - path (str or Path object): specify where file should be stored
            if not in current folder of calling script
    """
    if path is not None:
        output_name = str(path / output_name)
    with open(Path(output_name + '_analytically_mode_names_.csv'), 'w+') as file:
        for listitem in mode_names:
            file.write('%s\n' % listitem)


def load_data_from_exel_layered(
        material: str,
        file_name: str = 'A_Lamb',
        thickness: float = 1,
        save: bool = False
) -> tuple:
    """
    Load and convert frequency-velocity into wavenumber-frequency
        representation from DLR DispersionCalculator dispersion curves
        for LAYERED specimen. There is another function for single
        layer material.

    returns:
        - f: (ndarray) matrix of frequencies for each mode
        - v: (ndarray) matrix of phase speeds for each mode
        - mode_names: (list) list of all mode names
        - thickness: (float) Output of DC is f*d so we need the thickness
                to convert back to the real frequencies and wave numbers
        - output_name: (str) name of the related newly created output file

    args:
        - material: (string) - name of material and name of file where
            extracted data should be stored
        - file_name: (string or Path object) - name of the file of with
            stored dispersion data from DC without .xlsx file extension
        - save: (bool) - define if the converted dispersion data
            should be stored to .npy file
    """

    xlsx_file = Path(str(file_name) + '.xlsx')

    output_name = material + '_dispersion_data'

    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    min_column, max_column = 0, sheet.max_column

    # -- get Mode names
    mode_names = []
    locations = []
    for idx, column in enumerate(sheet.iter_cols(min_column, max_column)):
        if column[0].value[3:5] == 'fd':
            mode_names.append(column[0].value[0:2])
            locations.append(idx)

    # -- extract the data for each mode from document
    m = {}
    fs, vs = [], []
    num_elems = 0
    for mode_nr, col in enumerate(locations):
        for row in np.arange(2, sheet.max_row + 1):
            fs.append(sheet.cell(row=row, column=col + 1).value)
            vs.append(sheet.cell(row=row, column=col + 2).value)
        fs = [i for i in fs if isinstance(i, int) or isinstance(i, float)]
        vs = [i for i in vs if isinstance(i, int) or isinstance(i, float)]

        if len(fs) > 0 and len(vs) > 0:
            assert len(fs) == len(vs)
            m[mode_names[mode_nr]] = (list(zip(fs, vs)))
            if len(fs) > num_elems: num_elems = len(fs)
            fs, vs = [], []

    f = np.zeros((num_elems, len(m)))
    v = np.zeros((num_elems, len(m)))
    for idx_x, elem in enumerate(m):
        for idx_y, p in enumerate(m[elem]):
            f[idx_y, idx_x] = p[0] * 1E6
            v[idx_y, idx_x] = p[1] * 1E3

    f = np.divide(f, thickness)

    if save:
        save_dispersion_np_array(output_name, f, 'f')
        save_mode_names(output_name, mode_names)

    return f, v, mode_names, output_name


def convert_fc2kf(
        f: float,
        v: float,
        output_name: str,
        save: bool = False
) -> float:
    """
    Converts the frequency-velocity into wavenumber-frequency
        representation from DLR DispersionCalculator dispersion curves.

    returns:
        - k: (ndarray) matrix of wavenumbers for each mode. k is an
                array with shape num_elements x num_modes
        - f: (ndarray) matrix of frequencies for each mode
        - modenames (list) list of strings with associated mode names

    args:
        - f: (ndarray) frequency array for each mode
        - v: (ndarray) phase speed velocity for each mode
        - save: (bool) - define if the converted dispersion data
            should be stored to .npy file
    """

    # k = 2 * np.pi * np.divide(f, v)
    np.seterr(divide='ignore', invalid='ignore')
    k = np.divide(f, v)

    if save:
        save_dispersion_np_array(output_name, k, 'k')

    return k


def plot_analytical_dispersion_curves_single(
        material: str,
        k: float,
        f: float,
        mode_names: list,
        axis: bool = True,
        m_axis: list = None,
        path: str = None,
        save: bool = False,
        save_publication: bool = False
) -> None:
    """
    Plot dispersion curves from f, k, mn for a single material

    returns:
        - dispersion plot if specified

    args:
        - k: (ndarray) numpy array of wavenumbers
        - f: (ndarray) numpy array of frequencies
        - mode_names: (list) list of mode names
        - axis = True (bolean) - specify if axis should be rescaled or not
        - m_axis = None (list) - manually specified axis limits
        - path (str or Path object): specify where file should be stored
            if not in current folder of calling script
        - save - (boolean) - save plot or not
        - save_publication: bool = False : specifies if plot should be exported as pgf for
            publication in latex and if long title with meta data or cleaned title for publication
            should be included. If this is selected, make sure to comment in the rc commands in
            on top after matplotlib import!
    """
    fig = plt.figure(1, dpi=300)
    fig.set_size_inches(w=6, h=4)

    for row in range(k.shape[1]):
        if mode_names[row][0] == 'S':
            plt.plot(k[:, row], f[:, row], label=mode_names[row],
                     linewidth=.7)
        else:
            plt.plot(k[:, row], f[:, row], linestyle='--', label=mode_names[row],
                     linewidth=.7)

    if axis:
        plt.axis([0, 900, 0, 2.5e6])
        # plt.axis([0, 1200, 0, 3.5e6])
    elif m_axis is not None:
        plt.axis(m_axis)

    plt.title(f'Analytically obtained dispersion curves of ' + material)
    plt.xlabel(r'Wave number $k$ in $1/m$')
    plt.ylabel(r'Frequency $f$ in MHz')

    labels = ['A0','A1','A2','A3','A4','A5','A6','A7','A8','A9','S0','S1','S2','S3','S4','S5','S6','S7','S8','S9']
    plt.legend(labels, loc="lower center", bbox_to_anchor=(0.65, 0.01), ncol=4)  # ncol=4
    # plt.tight_layout()

    output_name = material + '_dispersion_data'
    if save:
        if save_publication:
            pub_out_name = f'Zirc-4_f_k_dispersion_plot.pgf'
            plt.savefig(pub_out_name[0:-3] + 'png', backend='pgf', format='png', dpi=200)
            plt.savefig(pub_out_name, backend='pgf', format='pgf')
        else:
            if path is not None:
                output_name = str(path / output_name)
            plt.savefig(Path(output_name + '.png'), bbox_inches='tight', dpi=300)
            plt.savefig(Path(output_name + '.eps'), format='eps', bbox_inches='tight', dpi=300)
    plt.show()


def get_label_names(material):
    """
    Obtain label names for plot from material

    args:
        - material (str): name of material used
            -> needs to be zy4cr for the Zirc4-Chrome specimen
    """
    if material != 'zy4cr':
        t_name = 'Alu'
        b_name = 'Tape'
        tb_name = 'AluTape'
    else:
        t_name = 'Cr'
        b_name = 'Zy4'
        tb_name = 'Zy4Cr'
    return t_name, b_name, tb_name


def get_plotting_layout(lo):
    """
    Specify which plotting layout should be used. Plotting layouts
    to choose from are:
        - koreck
        - analysis_sharp
        - analysis_thick
        - ...

    returns:
        - d (dict): dictionary with entries for color, linestyle, linewidth,
            alpha for the top layer, the bottom layer and the
            layered specimen

    args:
        - lo (str): specify which layout should be used. Must be
            within the set {'koreck', 'analysis_sharp', 'analysis_thick', ...}
    """

    if lo == 'koreck' or lo is None:
        t = {
            'c': 'm',
            'ls': '--',
            'd': (25, 10),
            'lw': .7,
            'a': .8,
        }
        b = {
            'c': 'k',
            'ls': '',
            'd': (None, None),
            'lw': .7,
            'a': .6,
        }
        tb = {
            'c': 'b',
            'ls': 'dashed',
            'd': (5, 5),
            'lw': .7,
            'a': .7,
        }
        dpi = 300
    elif lo == 'analysis_sharp':
        t = {
            'c': 'm',
            'ls': '-',
            'd': (None, None),
            'lw': .5,
            'a': .8,
        }
        b = {
            'c': 'k',
            'ls': '-',
            'd': (None, None),
            'lw': .5,
            'a': .6,
        }
        tb = {
            'c': 'b',
            'ls': 'solid',
            'd': (None, None),
            'lw': .5,
            'a': .7,
        }
        dpi = 600
    elif lo == 'analysis_thick':
        t = {
            'c': 'm',
            'ls': '-',
            'd': (None, None),
            'lw': .9,
            'a': .8,
        }
        b = {
            'c': 'k',
            'ls': '-',
            'd': (None, None),
            'lw': .9,
            'a': .6,
        }
        tb = {
            'c': 'b',
            'ls': 'solid',
            'd': (None, None),
            'lw': .9,
            'a': .7,
        }
        dpi = 600
    elif lo == 'analysis_max_sharp':
        t = {
            'c': 'r',
            'ls': '--',
            'd': (5, 3),
            'lw': .5,
            'a': .8,
        }
        b = {
            'c': 'y',
            'ls': '--',
            'd': (2, 1),
            'lw': .6,
            'a': .7,
        }
        tb = {
            'c': 'c',
            'ls': 'solid',
            'd': (None, None),
            'lw': .5,
            'a': .7,
        }
        dpi = 600
    elif lo == 'max_with_sim_old':
        t = {
            'c': 'r',
            'ls': '--',
            'd': (5, 3),
            'lw': .7,
            'a': .8,
        }
        b = {
            'c': 'y',
            'ls': '--',
            'd': (2, 1),
            'lw': .7,
            'a': .6,
        }
        tb = {
            'c': 'c',
            'ls': 'solid',
            'd': (None, None),
            'lw': .7,
            'a': .7,
        }
        dpi = 600
    else:
        raise(Exception('Specified plotting layout not valid!'
                        'Check for typos'))

    d = {
        't': t,
        'b': b,
        'tb': tb,
        'dpi': dpi
    }

    return d


def plot_analytical_dispersion_curves_layered(
        ka_t: float,
        fa_t: float,
        mn_t: float,
        ka_b: float,
        fa_b: float,
        mn_b: float,
        ka_tb: float,
        fa_tb: float,
        mn_tb: float,
        output_name: str,
        axis: bool = True,
        m_axis: list = None,
        material: str = 'zy4cr',
        path: str = None,
        layout: str = None,
        save: bool = False,
        show: bool = True
) -> None:
    """
    Plot analytical dispersion data for a layered specimen with two
    layers. Most probably, this will be Zirc-4/Cr, or Alu/Tape.

    returns:
        - plot of analytical dispersion curves

    args:
        - ka_t - analytically obtained wave numbers for top layer
        - fa_t - analytically obtained frequencies for top layer
        - mn_t - mode names for top layer
        - ka_b - analytically obtained wave numbers for bottom layer
        - fa_b - analytically obtained frequencies for bottom layer
        - mn_b - mode names for bottom layer
        - ka_tb - analytically obtained wave numbers for layered specimen (top + bottom)
        - fa_tb - analytically obtained frequencies for layered specimen (top + bottom)
        - mn_tb - mode names for layered specimen (top + bottom)
        - output_name - name of output file/plot caption
        - axis = True (boolean) - specify if axis should be rescaled or not
        - m_axis = None (list) - manually specified axis limits
        - path (str or Path object): specify where file should be stored
            if not in current folder of calling script
        - layout (str): specify which layout/plotting settings should be used
        - save = False (boolean) - flag of plot should be saved
        - show = True (boolean) - flag determining if plot should be shown
    """

    t_name, b_name, tb_name = get_label_names(material)

    d = get_plotting_layout(layout)

    plt.figure(1, dpi=d['dpi'])
    ax = plt.gca()

    for row in range(ka_t.shape[1]):  #[0, 1, 5, 6]:
        ax.plot(
            ka_t[:, row], fa_t[:, row],
            color=d['t']['c'],
            linestyle=d['t']['ls'],
            dashes=d['t']['d'],
            linewidth=d['t']['lw'],
            alpha=d['t']['a'],
            label=f'{t_name}-{mn_t[row]}'
        )
    for row in range(ka_b.shape[1]):
        ax.plot(
            ka_b[:, row],
            fa_b[:, row],
            color=d['b']['c'],
            linestyle=d['b']['ls'],
            dashes=d['b']['d'],
            linewidth=d['b']['lw'],
            alpha=d['b']['a'],
            label=f'{b_name}-{mn_b[row]}'
        )
    for row in range(ka_tb.shape[1]):
        ax.plot(
            ka_tb[:, row],
            fa_tb[:, row],
            color=d['tb']['c'],
            linestyle=d['tb']['ls'],
            dashes=d['tb']['d'],
            linewidth=d['tb']['lw'],
            alpha=d['tb']['a'],
            label=f'{tb_name}-{mn_tb[row]}'
        )

    if axis:
        plt.axis([0, 900, 0, 2.5e6])
    elif m_axis is not None:
        plt.axis(m_axis)

    plt.xlabel('Wavenumber k in 1/m')
    plt.ylabel('Frequency f in Hz')
    title = output_name + '_analytical_layered'
    plt.title(title)

    h, l = ax.get_legend_handles_labels()
    unique_labels = [lbl for i, lbl in enumerate(l) if i == 0
                     or (lbl[0:4] != l[i-1][0:4] and lbl[3] != 'S')]
    unique_idx = [l.index(lbl) for lbl in unique_labels]
    unique_labels = [lbl[0:lbl.find('-')] for lbl in unique_labels]
    plt.legend([h[idx] for idx in unique_idx], unique_labels)

    if save:
        if path is not None:
            output_name = str(path / output_name)
        plt.savefig(Path(output_name), bbox_inches='tight', pad_inches=0.1, dpi=600)
    if show:
        plt.show()


if __name__ == '__main__':
    a_path = Path("analytical_disp_curves\AluTape_dispersion_curves\Alutape")
    print(type(plot_analytical_dispersion_curves_single(material='AluTape')))
