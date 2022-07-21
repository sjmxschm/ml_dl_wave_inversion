import openpyxl

import numpy as np
import matplotlib.pyplot as plt

from pathlib import Path

'''
Script converts frequency-velocity into wave number-frequency representation. 

Input for this is the xlsx_file obtained from the DLR DispersionCalculator disperse software
which needs to be in the same folder as this script
'''


def convert_fc2kf(
        coating_thickness='0.05',
        plot=True,
        save=False,
):
    """
    Convert frequency-velocity into wavenumber-frequency representation from DLR DispersionCalculator dispersion curves.

    args:
        - coating_thickness: (str) - thickness of coating layer
        - ph_grp: (str) 'ph' or 'grp' - specify if group or phase velocity should be converted
        - plot: (bool) - define if converted dispersion curves should be plotted
        - save: (bool) - define if the converted dispersion data should be stored to .npy file
    """

    file_name = coating_thickness + '_Chrome' '/' '1_' + coating_thickness + '_Zy4Cr_dispersionplot'
    xlsx_file = Path(file_name + '.xlsx')

    output_name = coating_thickness + '_Chrome' '/' 'Zy4_3_Cr_' + coating_thickness + '_dispersion_data'

    print(xlsx_file)

    # -- read the active sheet with openpyxl:
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    min_column, max_column = 0, sheet.max_column

    # -- get Mode names
    mode_names = []
    locations = []
    for idx, column in enumerate(sheet.iter_cols(min_column, max_column)):
        if column[0].value[3:5] == 'fd':
            mode_names.append(column[0].value[0:2])  # [0] for testing - specify row with mode names here
            locations.append(idx)
    print(f'mode names = \n{mode_names}')

    # -- extract the data for each mode from document
    m = {}
    fs, vs = [], []
    num_elems = 0
    for mode_nr, col in enumerate(locations):
        for row in np.arange(2, sheet.max_row + 1):
            #import pdb; pdb.set_trace()
            fs.append(sheet.cell(row=row, column=col + 1).value)
            vs.append(sheet.cell(row=row, column=col + 2).value)
        fs = [i for i in fs if isinstance(i, int) or isinstance(i, float)]
        vs = [i for i in vs if isinstance(i, int) or isinstance(i, float)]

        if len(fs) > 0 and len(vs) > 0:
            assert len(fs) == len(vs)
            m[mode_names[mode_nr]] = (list(zip(fs, vs)))  # have a list of tuples for each mode in dict
            if len(fs) > num_elems: num_elems = len(fs)
            fs, vs = [], []

    f = np.zeros((num_elems, len(m)))
    v = np.zeros((num_elems, len(m)))
    for idx_x, elem in enumerate(m):
        for idx_y, p in enumerate(m[elem]):
            f[idx_y, idx_x] = p[0] * 1E6
            v[idx_y, idx_x] = p[1] * 1E3

    f = np.divide(f, float(coating_thickness))

    # k ist an array with shape num_elements x num_modes
    # k = 2 * np.pi * np.divide(f, v)
    k = np.divide(f, v)

    if plot:
        # -- plot dispersion curves for modes
        # range(k.shape[1]) - for all modes
        for row in range(k.shape[1]):
            plt.plot(k[:, row], f[:, row], label=mode_names[row])


        # plt.axis([0, 1500, 0, 2.5e6])
        plt.title(f'Analytically obtained dispersion curves\n{file_name}')
        plt.xlabel('Wavenumber k in 1/m')
        plt.ylabel('Frequency f in MHz')
        plt.legend()
        if save:
            plt.savefig(Path(output_name + '.png'), bbox_inches='tight', dpi=300)
        plt.show()

    if save:
        np.save(Path(output_name + '_analytically_k_'), k)
        np.save(Path(output_name + '_analytically_f_'), f)
        with open(Path(output_name + '_analytically_mode_names_.csv'), 'w+') as filehandle:
            for listitem in mode_names:
                filehandle.write('%s\n' % listitem)


if __name__ == '__main__':

    # thicknesses = ['0.05', '0.1', '0.2', '0.3', '0.4', '0.5', '0.6']

    thicknesses = ['0.03']

    for th in thicknesses:
        convert_fc2kf(
            coating_thickness=th,
            plot=True,
            save=True
        )
